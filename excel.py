import openpyxl as open
import xlwings as xw
arquivo = "G:/UnidadeN/PycharmProjects/calculadora_mercado.xlsx"
arquivo2 = 'G:/UnidadeN/PycharmProjects/test.xlsx'

def planilha_excel_xw():
    app = xw.App(visible=False)  # roda em segundo plano, pois sempre abre o excel

    wb = app.books.open(arquivo2)  # carrega o arquivo

    aba_ativa = wb.sheets['Precificação']  # Selecione a planilha desejada

    aba_ativa['D10'].value = 10.00  # Defina os valores nas células D10 para exemplo
    aba_ativa['D12'].value = 11.00  # Defina os valores nas células D12 para exemplo

    aba_ativa.range('F8').formula = '=SUM(D10:D12)'  # Defina a fórmula na célula E1 para somar as células D10 a D12
    aba_ativa.api.Calculate()  # atualiza o calculo da formula é como um f5

    print(f"Resultado da fórmula: {aba_ativa.range('F8').value}")

def carregar_dados_excel(operacao, vencimento, juros, volatilidade):
    app = xw.App(visible=False) # roda em segundo plano, pois sempre abre o excel

    wb = app.books.open(arquivo2) # carrega o arquivo

    aba_ativa = wb.sheets['Precificação'] # Selecione a planilha desejada

    coluna_c = aba_ativa.range('C1:C25') # define um range entre a coluna C

    for celula in coluna_c:
        if celula.value is not None:
            if celula.value == 'Ativo':
                ativo_v = aba_ativa['D6'].value = operacao.ativo.codigo
            if celula.value == 'Vencimento':
                vencimento_v = aba_ativa['D8'].value = vencimento
            if celula.value == 'Preço do ativo objeto':
                preco_ativo_v  = aba_ativa['D10'].value = operacao.ativo.ultimo_preco
            if celula.value == 'Strike':
                strike_v = aba_ativa['D12'].value = operacao.preco
            if celula.value == 'Juros':
                juros_v = aba_ativa['D14'].value = juros
            if celula.value == 'Volatilidade':
                volatilidade_v = aba_ativa['D16'].value = volatilidade

    aba_ativa.api.Calculate()

    preco_teorico_v = aba_ativa['D21'].value * 100
    delta_v = aba_ativa['D23'].value * 100
    lista_v = [ ativo_v,vencimento_v,preco_ativo_v,strike_v,juros_v,volatilidade_v,round(preco_teorico_v,2),round(delta_v,2)]
    print(lista_v)

    print("Ativo ")
    print(f"Preço teórico : {round(preco_teorico_v,2)}")
    print(f"Delta : {round(delta_v,2)}")

    wb.save()
    # Obtenha o resultado da fórmula na célula E1

    wb.close()
    app.quit()


operacao = ''
vencimento = ''
juros = ''
volatilidade = ''
carregar_dados_excel(operacao, vencimento, juros, volatilidade)


def planilha_exce():
    #carrega a planilha excel
    planilha = open.load_workbook(arquivo)
    #pega a celular principal ativa
    planilha._active_sheet_index = 0 #dessa forma também funciona, voce seta qual o indice das abas
    aba_ativa = planilha['Precificação']
    #imprime
    print(aba_ativa)
    ativo = aba_ativa['C6'].value
    vencimento = aba_ativa['C8'].value
    preco_ativo = aba_ativa['C10'].value
    strike  = aba_ativa['C12'].value
    juros = aba_ativa['C14'].value
    volatilidade = aba_ativa['C16'].value
    preco_teorico = aba_ativa['C21'].value
    delta = aba_ativa['C23'].value
    lista = [ativo,vencimento,preco_ativo,strike,juros,volatilidade,preco_teorico,delta]
    print(lista)

    '''
    Ativo
    Vencimento
    Preço do ativo objeto
    Strike
    Juros
    Volatilidade
    '''
    # for celula in aba_ativa.iter_rows(values_only=True):
    #     print(celula)
    #imprime as abas que existem na planilha
    for abas in planilha:
        print(abas)
#    linha = celula.row SE QUISER PEGAR A LINHA TODA.
# https://medium.com/data-hackers/como-manipular-planilhas-excel-com-o-python-6be8799f8dd7
    for celula in aba_ativa['C']:
        if celula.value == 'Ativo':
            ativo_v = aba_ativa['D6'] = 'BBAS3'
        if celula.value == 'Vencimento':
            vencimento_v = aba_ativa['D8'] = '21/10/2023'
        if celula.value == 'Preço do ativo objeto':
            preco_ativo_v  = aba_ativa['D10'] = 18.50
        if celula.value == 'Strike':
            strike_v = aba_ativa['D12'] = 19.20
        if celula.value == 'Juros':
            juros_v = aba_ativa['D14'] = 0.03
        if celula.value == 'Volatilidade':
            volatilidade_v = aba_ativa['D16'] = 0.4
    preco_teorico_v = aba_ativa['D21']
    delta_v = aba_ativa['D23']

    # soma = Tokenizer(f"=SOMA({aba_ativa['D10'].value};{aba_ativa['D12'].value})")
    # print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in soma.items))
    # print(soma)

    # soma2 = "=Soma(D10:D12)"
    # Translator(soma2, origin="E10")
    # aba_ativa.append([soma2])

    aba_ativa['F8'] = '=SUM(D10:D12)'

    aba_ativa['E11'] = 'BEM VINDO'
    aba_ativa['E12'] = 'A PLANILHA'
    aba_ativa['E13'] = 'NOVA '
    lista_v = [ ativo_v,vencimento_v,preco_ativo_v,strike_v,juros_v,volatilidade_v,preco_teorico_v,delta_v]
    print(lista_v)
    #recalcula no excel
    aba_ativa.calculation = open.workbook.calculation.Calculation(auto=True)

    planilha.save('G:/UnidadeN/PycharmProjects/test.xlsx', data_only=True)

    planilha2 = open.load_workbook("G:/UnidadeN/PycharmProjects/test.xlsx")
    aba_ativa2 = planilha2['Precificação']
    resultado_formula = aba_ativa2['F8'].value
    print(f'resultado : {resultado_formula}')
planilha_excel_xw()

